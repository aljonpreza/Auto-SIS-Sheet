from openpyxl import load_workbook
from collections import OrderedDict
import os
import sys
import requests
import json
import argparse

def parseOptions():
    # Command Line options
    if len(sys.argv) != 4:
        print 'usage: ./webscraperLexus <xlsxName> <sheetName> <fileName>'
        sys.exit(0)

    ret = {
        "xlsxName": sys.argv[1],
        "sheetName": sys.argv[2],
        "fileName": sys.argv[3]
    }

    return ret

def main():
    # Parses the options from the user
    opts = parseOptions()
    print opts

    try:
        wb = load_workbook(filename = opts["xlsxName"] + '.xlsx', data_only = True)
    except IOError:
        print "No xlsx file with that name."
        sys.exit(0)

    # Array of all sheets in spreadsheet
    wbSheets = wb.get_sheet_names()

    # Check whether sheet name given is valid
    if opts["sheetName"] in wbSheets:
        sheetIndex = wbSheets.index(opts["sheetName"])
        ws = wb[wbSheets[sheetIndex]]
    else:
        print "Name of sheet not found in xlsx."
        sys.exit(0)

    vehicleTypeYear = (ws['A1'].value).split()
    build = OrderedDict([
        ("vehicleType", vehicleTypeYear[0]),
        ("year", vehicleTypeYear[1]),
        ("vehicleStyle", ws['I3'].value),
        ("description", ws['A3'].value),
        ("packages", [])
    ])

    flag = 1 # Denotes end of available package iteration
    packageName = ""
    packagePrice = ""
    packageDetails = ""
    availColorRowStart = None # Used to mark the start of color selections
    # Iteration of Available Packages
    for row in ws.rows:
        for cell in row:
            if 'CHOICE' in str(cell.value):
                packageName = cell.value[len('CHOICE \"'):-len("\"")]
            if cell.column == 'F' and cell.value != None:
                if row[4].value == None:
                    packageDetails += row[6].value + "\n"
                else:
                    packageDetails += "%s ($%d)\n" % (row[6].value, row[9].value)
            if cell.column == 'I' and cell.value != None:
                # Gets total price of package, this also notes the end of a package.
                cellValue = cell.value[len('CHOICE \"'):-len("\" Total:")]
                if packageName == cellValue and row[9].value:
                    packagePrice = row[9].value
                    package = {
                        "packageName": packageName,
                        "packagePrice": packagePrice,
                        "packageDetails": packageDetails,
                        "exColors": [],
                        "intColors": []
                    }
                    build["packages"].append(package)
                    packageDetails = ""

            # End of all available packages, prints color combo starting row
            if cell.value == 'AVAILABLE COLOR COMBINATIONS':
                availColorRowStart = cell.row
                flag = 0
        if flag == 0:
            break

    # Get Colors
    exColors = []
    intColors = []
    numChoice = -1

    # Get interior and exterior colors of packages
    for row in ws[availColorRowStart:ws.max_row]:
        for cell in row:
            if cell.column == 'B' and cell.value != None and cell.value != 'CHOICE(S)':
                numChoice += 1
                currentChoice = cell.value
            # Inserts color to exterior and interior selection array without duplicates
            if cell.column == 'G' and cell.value != None and cell.value != 'COLOR':
                tempExt = {"colorName": cell.value, "isApplicable": True}
                if tempExt not in build['packages'][numChoice]['exColors']:
                    build['packages'][numChoice]['exColors'].append(tempExt)
            if cell.column == 'J' and cell.value != None and cell.value != 'COLOR':
                tempInt = {"colorName": cell.value, "isApplicable": True}
                if tempInt not in build['packages'][numChoice]['intColors']:
                    build['packages'][numChoice]['intColors'].append(tempInt)

    with open("./scripts/json/" + opts["fileName"] + ".json", 'w') as outfile:
        json.dump(build, outfile, indent = 2)
    print "\nJSON scraped into " + opts["fileName"] + ".json"

if __name__ == '__main__':
    main()
